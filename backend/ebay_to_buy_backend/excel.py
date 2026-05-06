import datetime as dt
import io
from dataclasses import dataclass

import asyncpg
from openpyxl import Workbook, load_workbook


def _excel_safe(value):
    """Excel не поддерживает timezone-aware datetime."""
    if isinstance(value, dt.datetime) and value.tzinfo is not None:
        return value.astimezone(dt.timezone.utc).replace(tzinfo=None)
    return value

# [[Excel import#columns]] — точные имена колонок, регистр обязателен.
_REQUIRED_HEADERS = ("smart_part_id", "target_qty", "is_active")

# [[purchase_overview#columns]] — порядок и имена.
OVERVIEW_COLUMNS = (
    "smart_part_id",
    "smart_name",
    "articles_text",
    "target_qty",
    "stock_total_qty",
    "need_qty",
    "is_need",
    "is_active",
    "active_ebay_count",
    "active_ebay_item_numbers",
    "active_ebay_comments",
    "ended_ebay_count",
    "ended_ebay_item_numbers",
    "ended_ebay_comments",
    "created_at",
    "updated_at",
)


@dataclass
class ImportResult:
    inserted: int
    updated: int


def _parse_is_active(raw: object, row_number: int) -> bool | None:
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)) and raw in (0, 1):
        return bool(raw)
    if isinstance(raw, str):
        s = raw.strip().lower()
        if s in ("true", "1", "да"):
            return True
        if s in ("false", "0", "нет"):
            return False
    raise ValueError(
        f"строка {row_number}: невалидное значение is_active: {raw!r} "
        f"(допустимо: пусто, true/false, 1/0, да/нет)"
    )


def _parse_target_qty(raw: object, row_number: int) -> int:
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        raise ValueError(f"строка {row_number}: target_qty обязателен")
    if isinstance(raw, bool):  # bool — подкласс int, отделяем явно
        raise ValueError(f"строка {row_number}: target_qty должен быть целым числом")
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    if isinstance(raw, str) and raw.strip().lstrip("-").isdigit():
        return int(raw.strip())
    raise ValueError(f"строка {row_number}: target_qty должен быть целым числом, получено {raw!r}")


def _parse_smart_part_id(raw: object, row_number: int) -> str:
    if raw is None or not isinstance(raw, str) or not raw.strip():
        raise ValueError(f"строка {row_number}: smart_part_id обязателен")
    return raw.strip()


# [[Excel import#rules]]: upsert; обновляются target_qty и is_active;
# отсутствие в файле не удаляет; пустой is_active для существующей не трогает.
_IMPORT_SQL = """
INSERT INTO purchase_targets (smart_part_id, target_qty, is_active)
VALUES ($1, $2, COALESCE($3, true))
ON CONFLICT (smart_part_id) DO UPDATE SET
    target_qty = EXCLUDED.target_qty,
    is_active  = COALESCE($3, purchase_targets.is_active)
RETURNING (xmax = 0) AS inserted
"""


async def import_targets(pool: asyncpg.Pool, file_bytes: bytes) -> ImportResult:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        raise ValueError("xlsx пустой: нет строки-заголовка")

    header_index: dict[str, int] = {}
    for idx, name in enumerate(header):
        if isinstance(name, str):
            header_index[name] = idx
    missing = [c for c in _REQUIRED_HEADERS if c not in header_index]
    if missing:
        raise ValueError(
            f"xlsx: отсутствуют обязательные колонки: {missing}; "
            f"требуется {list(_REQUIRED_HEADERS)} по [[Excel import#columns]]"
        )
    i_smart = header_index["smart_part_id"]
    i_qty = header_index["target_qty"]
    i_active = header_index["is_active"]

    inserted = 0
    updated = 0
    header_len = len(header)
    async with pool.acquire() as conn:
        async with conn.transaction():
            for row_number, row in enumerate(rows_iter, start=2):
                if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                    continue  # пустые строки в конце файла
                # openpyxl read_only обрезает trailing None — паддим обратно.
                padded = list(row) + [None] * (header_len - len(row))
                smart_part_id = _parse_smart_part_id(padded[i_smart], row_number)
                target_qty = _parse_target_qty(padded[i_qty], row_number)
                is_active = _parse_is_active(padded[i_active], row_number)
                rec = await conn.fetchrow(_IMPORT_SQL, smart_part_id, target_qty, is_active)
                if rec["inserted"]:
                    inserted += 1
                else:
                    updated += 1
    wb.close()
    return ImportResult(inserted=inserted, updated=updated)


# [[Excel export#filters]] — фильтры на purchase_overview.
def _build_overview_query(
    is_need: bool | None,
    is_active: bool | None,
    has_active_ebay: bool | None,
) -> tuple[str, list]:
    where: list[str] = []
    params: list = []
    if is_need is not None:
        params.append(is_need)
        where.append(f"is_need = ${len(params)}")
    if is_active is not None:
        params.append(is_active)
        where.append(f"is_active = ${len(params)}")
    if has_active_ebay is not None:
        if has_active_ebay:
            where.append("active_ebay_count > 0")
        else:
            where.append("active_ebay_count = 0")
    sql = "SELECT * FROM purchase_overview"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY smart_part_id"
    return sql, params


async def export_overview(
    pool: asyncpg.Pool,
    *,
    is_need: bool | None = None,
    is_active: bool | None = None,
    has_active_ebay: bool | None = None,
    explode_articles: bool = False,
    explode_active_ebay: bool = False,
    explode_ended_ebay: bool = False,
) -> bytes:
    sql, params = _build_overview_query(is_need, is_active, has_active_ebay)
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql, *params)
    rows = [dict(r) for r in rows]

    # Раскладки — считаем максимальное количество элементов по выборке.
    def split_by_comma(value: str | None) -> list[str]:
        if not value:
            return []
        return [s.strip() for s in value.split(", ")]

    def split_by_newline(value: str | None) -> list[str]:
        if not value:
            return []
        return value.split("\n")

    articles_lists = [split_by_comma(r["articles_text"]) for r in rows]
    active_num_lists = [split_by_comma(r["active_ebay_item_numbers"]) for r in rows]
    active_com_lists = [split_by_newline(r["active_ebay_comments"]) for r in rows]
    ended_num_lists = [split_by_comma(r["ended_ebay_item_numbers"]) for r in rows]
    ended_com_lists = [split_by_newline(r["ended_ebay_comments"]) for r in rows]
    max_articles = max((len(x) for x in articles_lists), default=0)
    max_active = max(
        (max(len(n), len(c)) for n, c in zip(active_num_lists, active_com_lists)),
        default=0,
    )
    max_ended = max(
        (max(len(n), len(c)) for n, c in zip(ended_num_lists, ended_com_lists)),
        default=0,
    )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("overview")

    # Шапка.
    header: list[str] = []
    for col in OVERVIEW_COLUMNS:
        if col == "articles_text" and explode_articles:
            for i in range(1, max_articles + 1):
                header.append(f"article_{i}")
        elif col == "active_ebay_item_numbers" and explode_active_ebay:
            for i in range(1, max_active + 1):
                header.append(f"active_ebay_item_number_{i}")
                header.append(f"active_ebay_comment_{i}")
        elif col == "active_ebay_comments" and explode_active_ebay:
            continue  # склеено с active_ebay_item_numbers выше
        elif col == "ended_ebay_item_numbers" and explode_ended_ebay:
            for i in range(1, max_ended + 1):
                header.append(f"ended_ebay_item_number_{i}")
                header.append(f"ended_ebay_comment_{i}")
        elif col == "ended_ebay_comments" and explode_ended_ebay:
            continue
        else:
            header.append(col)
    ws.append(header)

    for idx, r in enumerate(rows):
        line: list = []
        articles = articles_lists[idx]
        active_nums = active_num_lists[idx]
        active_coms = active_com_lists[idx]
        ended_nums = ended_num_lists[idx]
        ended_coms = ended_com_lists[idx]
        for col in OVERVIEW_COLUMNS:
            if col == "articles_text" and explode_articles:
                for i in range(max_articles):
                    line.append(articles[i] if i < len(articles) else None)
            elif col == "active_ebay_item_numbers" and explode_active_ebay:
                for i in range(max_active):
                    line.append(active_nums[i] if i < len(active_nums) else None)
                    line.append(active_coms[i] if i < len(active_coms) else None)
            elif col == "active_ebay_comments" and explode_active_ebay:
                continue
            elif col == "ended_ebay_item_numbers" and explode_ended_ebay:
                for i in range(max_ended):
                    line.append(ended_nums[i] if i < len(ended_nums) else None)
                    line.append(ended_coms[i] if i < len(ended_coms) else None)
            elif col == "ended_ebay_comments" and explode_ended_ebay:
                continue
            else:
                line.append(_excel_safe(r[col]))
        ws.append(line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
